"""
Export utilities: CSV bytes, Excel bytes, clipboard string.
"""

import io
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from schema import CURRENCY_FIELDS, PERCENT_FIELDS


def to_csv(df: pd.DataFrame) -> bytes:
    """Return UTF-8 CSV bytes of the DataFrame."""
    return df.to_csv(index=False).encode("utf-8")


def to_excel(df: pd.DataFrame) -> bytes:
    """
    Return Excel bytes with:
    - Frozen header row
    - Bold header with blue fill
    - Currency formatting for dollar columns
    - Percentage formatting for cap rate columns
    - Auto-width columns (capped at 40)
    """
    buffer = io.BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Comps")
        ws = writer.sheets["Comps"]

        # Header styling
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        thin = Side(style="thin", color="CCCCCC")
        border = Border(bottom=thin)

        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", wrap_text=False)
            cell.border = border

        # Freeze header row
        ws.freeze_panes = "A2"

        # Column-level number formatting
        for col_idx, col_name in enumerate(df.columns, start=1):
            col_letter = get_column_letter(col_idx)
            if col_name in CURRENCY_FIELDS:
                fmt = '#,##0'
            elif col_name in PERCENT_FIELDS:
                fmt = '0.00%'
            else:
                fmt = None

            if fmt:
                for row_idx in range(2, ws.max_row + 1):
                    ws[f"{col_letter}{row_idx}"].number_format = fmt

        # Auto-width
        for col_idx, col_name in enumerate(df.columns, start=1):
            col_letter = get_column_letter(col_idx)
            max_len = max(
                len(str(col_name)),
                *[len(str(ws.cell(row=r, column=col_idx).value or ""))
                  for r in range(2, min(ws.max_row + 1, 102))]
            )
            ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

    return buffer.getvalue()


def to_clipboard_tsv(df: pd.DataFrame) -> str:
    """
    Return tab-separated string (with header) for pasting directly into Excel.
    pyperclip.copy() this result to put it on the system clipboard.
    """
    return df.to_csv(sep="\t", index=False)
